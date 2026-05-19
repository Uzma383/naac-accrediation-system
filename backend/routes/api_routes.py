import os
import uuid
from flask import Blueprint, jsonify, request, session, current_app, send_file
from werkzeug.utils import secure_filename
from models.models import *
from extensions import db, bcrypt
import pandas as pd
from datetime import datetime, date
from decimal import Decimal

api_bp = Blueprint('api_bp', __name__, url_prefix='/api')

# --- Auth Endpoints ---

@api_bp.route('/register', methods=['POST'])
def api_register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'Admin')
    department = data.get('department', '')
    user_identifier = data.get('user_identifier', '')
    year = data.get('year', '')

    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({"success": False, "error": "Username already exists"}), 400

    hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
    new_user = User(
        username=username, 
        password=hashed_pw, 
        role=role, 
        department=department,
        user_identifier=user_identifier,
        year=year
    )
    db.session.add(new_user)
    db.session.commit()
    
    return jsonify({"success": True, "message": "Account created"})


@api_bp.route('/login', methods=['POST'])
def api_login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    academic_year = data.get('academic_year')
    
    if not username or not password:
        return jsonify({"success": False, "error": "Username and password required"}), 400
        
    user = User.query.filter_by(username=username).first()
    if user and bcrypt.check_password_hash(user.password, password):
        session['user_id'] = user.id
        session['academic_year'] = academic_year
        
        return jsonify({
            "success": True, 
            "user": {
                "id": user.id, 
                "username": user.username, 
                "role": user.role, 
                "department": user.department,
                "academic_year": academic_year or user.year,
                "program": "MCA",
                "programCode": "515124110"
            }
        })
    return jsonify({"success": False, "error": "Invalid username or password"}), 401


# --- Metadata Endpoints ---

@api_bp.route('/teachers', methods=['GET'])
def get_teachers():
    teachers = Teacher.query.all()
    return jsonify([{"id": t.id, "name": t.name} for t in teachers])

@api_bp.route('/departments', methods=['GET'])
def get_departments():
    progs = Program.query.all()
    res, seen = [], set()
    for p in progs:
        dept = p.department or p.program_name
        if dept not in seen:
            seen.add(dept)
            res.append({"id": p.id, "code": dept, "programCode": p.program_code, "programName": p.program_name})
    if not res:
        res = [{"id": 1, "code": "MCA", "programCode": "515124110", "programName": "MCA"}]
    return jsonify(res)

@api_bp.route('/semesters', methods=['GET'])
def get_semesters():
    recs = SemesterLookup.query.all()
    if not recs:
        # Default seed
        defaults = ["FYMCA-SEM-I (MCA)", "FYMCA-SEM-II (MCA)", "SYMCA-SEM-III (MCA)", "SYMCA-SEM-IV (MCA)"]
        for d in defaults: db.session.add(SemesterLookup(value=d))
        db.session.commit()
        recs = SemesterLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/courses', methods=['GET'])
def get_courses():
    """
    GET /api/courses          → all courses in SemesterCourseLookup
    GET /api/courses?sem=X    → courses for that semester only
    """
    sem = request.args.get('sem')
    q = SemesterCourseLookup.query
    if sem:
        q = q.filter(SemesterCourseLookup.semester == sem)
    rows = q.order_by(SemesterCourseLookup.course_code).all()
    return jsonify([{"code": r.course_code, "name": r.course_name, "semester": r.semester} for r in rows])

@api_bp.route('/courses', methods=['POST'])
def add_course():
    """
    POST /api/courses  { semester, course_code, course_name }
    Adds a new course to SemesterCourseLookup with normalisation + uniqueness check.
    """
    data = request.json or {}
    sem    = (data.get('semester') or '').strip()
    c_code = (data.get('course_code') or '').strip().upper()
    c_name = (data.get('course_name') or '').strip()

    if not sem or not c_code or not c_name:
        return jsonify({"success": False, "error": "semester, course_code and course_name are required"}), 400

    existing = SemesterCourseLookup.query.filter_by(semester=sem, course_code=c_code).first()
    if existing:
        return jsonify({"success": False, "error": f"Course {c_code} already exists in {sem}"}), 400

    try:
        new_c = SemesterCourseLookup(semester=sem, course_code=c_code, course_name=c_name)
        db.session.add(new_c)
        db.session.commit()
        return jsonify({"success": True, "course": {"code": c_code, "name": c_name, "semester": sem}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400


@api_bp.route('/electives', methods=['GET'])
def get_electives():
    # Attempting to load by 310916% or elective
    courses = Course.query.filter((Course.course_code.like('310916%')) | (Course.course_name.ilike('%elective%'))).all()
    return jsonify([{"id": c.id, "code": c.course_code, "name": c.course_name} for c in courses])

@api_bp.route('/program-codes', methods=['GET'])
def get_program_codes():
    progs = Program.query.all()
    return jsonify([{"code": p.program_code, "label": p.program_code} for p in progs])

@api_bp.route('/course-types', methods=['GET'])
def get_course_types():
    recs = CourseTypeLookup.query.all()
    if not recs:
        defaults = ["PBL (Project Based Learning)", "Major Project"]
        for d in defaults: db.session.add(CourseTypeLookup(value=d))
        db.session.commit()
        recs = CourseTypeLookup.query.all()
    # Note: legacy code returned courseCode. Keeping the structure.
    return jsonify([{"value": x.value, "label": x.value, "courseCode": ""} for x in recs])

@api_bp.route('/academic-years', methods=['GET'])
def get_academic_years():
    recs = AcademicYearLookup.query.all()
    if not recs:
        defaults = ["2020-21", "2021-22", "2022-23", "2023-24", "2024-25", "2025-26", "2026-27"]
        for d in defaults: db.session.add(AcademicYearLookup(value=d))
        db.session.commit()
        recs = AcademicYearLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/programmes', methods=['GET'])
def get_programmes():
    progs = Program.query.all()
    if not progs:
        return jsonify([
            {"code": "MCA",  "name": "Master of Computer Applications (MCA)"},
            {"code": "MBA",  "name": "Master of Business Administration (MBA)"}
        ])
    return jsonify([{"code": p.program_code, "name": p.program_name} for p in progs])

@api_bp.route('/reserved-categories', methods=['GET'])
def get_reserved_categories():
    recs = ReservedCategoryLookup.query.all()
    if not recs:
        defaults = ["SC", "ST", "OBC", "Divyangjan", "Gen-EWS", "Others"]
        for d in defaults: db.session.add(ReservedCategoryLookup(value=d))
        db.session.commit()
        recs = ReservedCategoryLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/library-resources', methods=['GET'])
def get_library_resources():
    recs = LibraryResourceLookup.query.all()
    if not recs:
        defaults = ["Books", "Journals", "e-Journals", "e-Books", "e-ShodhSindhu", "Shodhganga", "Databases"]
        for d in defaults: db.session.add(LibraryResourceLookup(value=d))
        db.session.commit()
        recs = LibraryResourceLookup.query.all()
    return jsonify([x.value for x in recs])

@api_bp.route('/qualifying-exams', methods=['GET'])
def get_qualifying_exams():
    recs = QualifyingExamLookup.query.all()
    if not recs:
        defaults = ["NET", "SLET", "GATE", "GMAT", "CAT", "GRE", "JAM", "IELTS", "TOEFL", "Civil Services", "State government examinations", "Other examinations"]
        for d in defaults: db.session.add(QualifyingExamLookup(value=d))
        db.session.commit()
        recs = QualifyingExamLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/event-levels', methods=['GET'])
def get_event_levels():
    recs = EventLevelLookup.query.all()
    if not recs:
        defaults = ["Inter-university", "State", "National", "International", "District"]
        for d in defaults: db.session.add(EventLevelLookup(value=d))
        db.session.commit()
        recs = EventLevelLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

@api_bp.route('/award-categories', methods=['GET'])
def get_award_categories():
    recs = AwardCategoryLookup.query.all()
    if not recs:
        defaults = ["Individual", "Team"]
        for d in defaults: db.session.add(AwardCategoryLookup(value=d))
        db.session.commit()
        recs = AwardCategoryLookup.query.all()
    return jsonify([{"value": x.value, "label": x.value} for x in recs])

# --- Generic POST Endpoint to add new lookup items ---
LOOKUP_MODELS = {
    "semesters": SemesterLookup,
    "course-types": CourseTypeLookup,
    "academic-years": AcademicYearLookup,
    "reserved-categories": ReservedCategoryLookup,
    "library-resources": LibraryResourceLookup,
    "qualifying-exams": QualifyingExamLookup,
    "event-levels": EventLevelLookup,
    "award-categories": AwardCategoryLookup,
    "departments": DepartmentLookup,
    "designations": DesignationLookup,
    "highest-degrees": HighestDegreeLookup,
    "appointment-types": AppointmentTypeLookup,
    "funding-agencies": FundingAgencyLookup,
    "levels": LevelLookup,
    "team-individual": TeamIndividualLookup,
    "egovernance-areas": EGovernanceAreaLookup
}

# --- Dedicated Getters for new Lookups ---

@api_bp.route('/get-lookups/<lookup_key>', methods=['GET'])
def get_lookup_values(lookup_key):
    model = LOOKUP_MODELS.get(lookup_key)
    if not model: return jsonify([])
    recs = model.query.all()
    # Handle different models if needed, but most use .value
    return jsonify([{"id": r.id, "value": r.value, "label": r.value} for r in recs])

@api_bp.route('/lookups/<lookup_key>', methods=['POST'])
def add_lookup_value(lookup_key):
    model = LOOKUP_MODELS.get(lookup_key)
    if not model:
        return jsonify({"success": False, "error": "Invalid lookup key"}), 404
        
    data = request.json
    new_val = data.get('value')
    if not new_val:
        return jsonify({"success": False, "error": "Value required"}), 400
        
    existing = model.query.filter_by(value=new_val).first()
    if not existing:
        try:
            new_item = model(value=new_val)
            db.session.add(new_item)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 400
            
    return jsonify({"success": True, "message": f"Added {new_val}"})

# --- Generic Criteria Router ---

CRITERIA_MODELS = {
    "1_1": C11Courses, "1_1_3": C113TeacherBodies, "1_2_1": C121CBCS, "1_2_2": C122Addon, 
    "1_3_2": C132Experiential, "1_3_3": C133Projects, 
    "2_1": C21Students, "2_2": C22ReservedCategorySeats,
    "2_1_1": C211Enrolment, "2_1_2": C212Reservation, "2_3": C23OutgoingStudents,
    "2_3_3": C233MentorRatio, "2_4_1": C241Teachers, "2_4_2": C242TeacherPhD, "2_6_3": C263PassPercentage,
    "3_1": C3FullTimeTeachers, "3_2": C3SanctionedPosts, "3_1_1_2": C3ResearchProjects, "3_1_3": C313Events,
    "3_2_1": C321Papers, "3_2_2": C322Books, "3_3_2": C332ExtensionAwards, "3_3_3_4": C333Outreach,
    "3_4_1": C341Collaborations, "3_4_2": C342MoUs,
    "4_1_3": C413ICTRooms, "4_1_4": C4Expenditure, "4_2_2": C42Library,
    "5_1_1": C511Scholarships, "5_1_3": C513SkillInitiatives, "5_1_4": C514CompetitiveExams,
    "5_2_1": C521Placements, "5_2_2": C522HigherEd, "5_2_3": C523QualifyingExams,
    "5_3_1": C531SportsAwards, "5_3_3": C533SportsEvents,
    "6_2_3": C623EGovernance, "6_3_2": C632TeacherFinancial, "6_3_3": C633StaffTraining,
    "6_3_4": C634TeacherFDP, "6_4_2": C642NonGovGrants, "6_5_3": C653QualityInitiatives
}

from decimal import Decimal
from datetime import date

def resolve_student_id(model, data):
    if not hasattr(model, 'student_id'):
        return None
    s_name = data.get('studentName') or data.get('student_name')
    if not s_name:
        return None
    student = Student.query.filter_by(name=s_name).first()
    if not student:
        p_code = data.get('programCode') or data.get('program') or data.get('program_graduated')
        reg_no = data.get('registration_no') or data.get('reg_no') or data.get('enrollmentNumber') or data.get('registration_number')
        
        # Unique/Nullable constraint handling for enrollment_number
        if not reg_no:
            import uuid
            reg_no = f"TEMP-{uuid.uuid4().hex[:8].upper()}"
            
        prog = None
        if p_code:
            prog = Program.query.filter((Program.program_code == p_code) | (Program.program_name == p_code)).first()
            
        student = Student(
            name=s_name,
            enrollment_number=reg_no,
            program_id=prog.id if prog else None,
            category=data.get('category')
        )
        db.session.add(student)
        db.session.flush()
    return student.id

def resolve_teacher_id(model, data):
    teacher_val = data.get('teacher_id') or data.get('teacherName') or data.get('pi_name') or data.get('pi_id') or data.get('name_of_teacher') or data.get('name_of_teacher_who_attended')
    if not teacher_val:
        return None
    try:
        t_id = int(teacher_val)
        t = Teacher.query.get(t_id)
        if t:
            return t.id
    except ValueError:
        pass
    t = Teacher.query.filter_by(name=str(teacher_val).strip()).first()
    if not t:
        joining_date_val = None
        j_date_str = data.get('date_of_joining') or data.get('joining_date')
        if j_date_str:
            try:
                joining_date_val = datetime.strptime(j_date_str, '%d-%m-%Y').date()
            except ValueError:
                try:
                    joining_date_val = datetime.strptime(j_date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
        
        t = Teacher(
            name=str(teacher_val).strip(),
            aadhar_or_id=data.get('id_number_aadhar') or data.get('aadhar_or_id') or data.get('pan'),
            email=data.get('email'),
            gender=data.get('gender'),
            designation=data.get('designation'),
            joining_date=joining_date_val,
            highest_degree=data.get('highest_degree'),
            degree_year=int(data['degree_year']) if data.get('degree_year') and str(data['degree_year']).isdigit() else None
        )
        db.session.add(t)
        db.session.flush()
    else:
        updated = False
        if data.get('email') and t.email != data.get('email'):
            t.email = data.get('email')
            updated = True
        if data.get('gender') and t.gender != data.get('gender'):
            t.gender = data.get('gender')
            updated = True
        if data.get('designation') and t.designation != data.get('designation'):
            t.designation = data.get('designation')
            updated = True
        if data.get('id_number_aadhar') and t.aadhar_or_id != data.get('id_number_aadhar'):
            t.aadhar_or_id = data.get('id_number_aadhar')
            updated = True
        if updated:
            db.session.flush()
            
    return t.id

def to_dict(rec):
    d = {}
    for col in rec.__table__.columns:
        if col.name in ['created_at', 'updated_at', 'created_by_id', 'updated_by_id']:
            continue
        val = getattr(rec, col.name)
        if isinstance(val, date): val = val.isoformat()
        if isinstance(val, Decimal): val = float(val)
        if isinstance(val, datetime): val = val.isoformat()
        d[col.name] = val
        
        # UI Aliasing
        if col.name == 'academic_year' or col.name == 'year_of_offering':
            d['year'] = val
        if col.name == 'status_of_implementation':
            d['cbcsStatus'] = "Yes" if val else "No"
        if col.name == 'year_of_implementation':
            d['cbcsYear'] = val
        if col.name == 'institution_joined':
            d['inst_joined'] = val
        if col.name == 'program_admitted':
            d['prog_joined'] = val
        if col.name == 'registration_number':
            d['registration_no'] = val
        if col.name == 'exam_type':
            d['exam_qualified'] = val
        if col.name == 'program_graduated':
            d['program'] = val
        if col.name == 'supporting_document':
            d['upload_supporting_document'] = val
        if col.name.startswith('earmarked_'):
            d[col.name.replace('earmarked_', 'ear_')] = val
        if col.name.startswith('admitted_'):
            d[col.name.replace('admitted_', 'adm_')] = val

    # Lookup related names for the UI tables
    if hasattr(rec, 'course_id') and rec.course_id:
        c = Course.query.get(rec.course_id)
        if c: d['courseCode'] = c.course_code; d['courseName'] = c.course_name
    if hasattr(rec, 'program_id') and rec.program_id:
        p = Program.query.get(rec.program_id)
        if p: d['programCode'] = p.program_code; d['programName'] = p.program_name
    if hasattr(rec, 'teacher_id') and rec.teacher_id:
        t = Teacher.query.get(rec.teacher_id)
        if t:
            d['teacherName'] = t.name
            d['teacher_id'] = t.id
            d['id_number_aadhar'] = t.aadhar_or_id
            d['email'] = t.email
            d['gender'] = t.gender
            d['designation'] = t.designation
            d['date_of_joining'] = t.joining_date.strftime('%d-%m-%Y') if isinstance(t.joining_date, (date, datetime)) else str(t.joining_date) if t.joining_date else ''
    if hasattr(rec, 'pi_id') and rec.pi_id:
        t = Teacher.query.get(rec.pi_id)
        if t:
            d['pi_name'] = t.name
            d['pi_id'] = t.id
            d['teacherName'] = t.name
            d['id_number_aadhar'] = t.aadhar_or_id
            d['email'] = t.email
            d['gender'] = t.gender
            d['designation'] = t.designation
            d['date_of_joining'] = t.joining_date.strftime('%d-%m-%Y') if isinstance(t.joining_date, (date, datetime)) else str(t.joining_date) if t.joining_date else ''
    if hasattr(rec, 'student_id') and rec.student_id:
        s = Student.query.get(rec.student_id)
        if s:
            d['enrollmentNumber'] = s.enrollment_number
            d['studentName'] = s.name
            d['student_name'] = s.name

    if hasattr(rec, 'team_or_individual'):
        d['team_individual'] = rec.team_or_individual
    if hasattr(rec, 'level'):
        d['event_level'] = rec.level
    if hasattr(rec, 'activity_type'):
        d['event_name'] = rec.activity_type

    # ---- Criteria 6 aliases ----
    cls = rec.__class__.__name__
    if cls == 'C623EGovernance':
        d['areas_of_e_governance'] = d.get('area', '')
        vendor_parts = [d.get('vendor_name') or '', d.get('vendor_contact') or '']
        d['name_of_vendor_with_contact_details'] = ' '.join(p for p in vendor_parts if p).strip()
    elif cls == 'C632TeacherFinancial':
        d['name_of_teacher'] = d.get('teacherName', '')
        d['name_of_conference_workshop'] = d.get('conference_name', '')
        d['name_of_professional_body'] = d.get('professional_body', '')
        d['amount_of_support'] = str(d.get('amount', '')) if d.get('amount') is not None else ''
    elif cls == 'C633StaffTraining':
        if not d.get('dates_from_to'):
            df = d.get('date_from', '') or ''
            dt = d.get('date_to', '') or ''
            d['dates_from_to'] = f"{df} to {dt}" if df and dt else df or dt or ''
        d['title_of_professional_development_program'] = d.get('teaching_program_title', '')
        d['title_of_administrative_training_program'] = d.get('non_teaching_program_title', '')
        d['no_of_participants'] = d.get('participant_count', '')
    elif cls == 'C634TeacherFDP':
        d['name_of_teacher_who_attended'] = d.get('teacherName', '')
        d['title_of_the_program'] = d.get('program_title', '')
        if not d.get('duration_from_to'):
            df = d.get('duration_from', '') or ''
            dt = d.get('duration_to', '') or ''
            d['duration_from_to'] = f"{df} to {dt}" if df and dt else df or dt or ''
    elif cls == 'C642NonGovGrants':
        d['name_of_non_government_funding_agencies_individuals'] = d.get('agency_name', '')
        d['purpose_of_the_grant'] = d.get('purpose', '')
        d['funds_grants_received_inr_in_lakhs'] = str(d.get('amount_received', '')) if d.get('amount_received') is not None else ''
        d['link_to_audited_statement'] = d.get('audited_statement_link', '')
    elif cls == 'C653QualityInitiatives':
        d['conferences_seminars_workshops_on_quality'] = d.get('conferences_conducted', '')
        d['academic_administrative_audit_aaa'] = d.get('aaa_status', '')
        d['participation_in_nirf'] = d.get('nirf_status', '')
        d['nba_or_other_certification'] = d.get('nba_certification', '')
        d['collaborative_quality_initiatives'] = d.get('collaborative_initiatives', '')
        d['orientation_programme_on_quality_issues'] = d.get('orientation_program', '')

    return d

@api_bp.route('/records/<criterion>', methods=['GET'])
def get_records(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify([])
    
    if criterion == '1_1':
        # Enrich with semester name (stored as programName in C11Courses via program relation)
        records = C11Courses.query.all()
        result = []
        for r in records:
            d = to_dict(r)
            # Also resolve semester from SemesterCourseLookup using course_code
            if r.course_id:
                c = Course.query.get(r.course_id)
                if c:
                    scl = SemesterCourseLookup.query.filter_by(course_code=c.course_code).first()
                    if scl: d['programName'] = scl.semester
            result.append(d)
        return jsonify(result)
    
    return jsonify([to_dict(r) for r in model.query.all()])

# ---- Dedicated POST for Criterion 1.1 ----
@api_bp.route('/records/1_1', methods=['POST'])
def add_record_1_1():
    """
    Handles Criterion 1.1 form submission.
    UI sends: { department, programCode, programName (=semester), courseCode, courseName, year }
    Maps to C11Courses: { academic_year, program_id, course_id }
    """
    data = request.json or {}
    
    semester    = data.get('programName', '').strip()  # "FYMCA-SEM-I", etc.
    program_code= data.get('programCode', '').strip()
    course_code = data.get('courseCode', '').strip()
    year        = str(data.get('year', '')).strip()
    department  = data.get('department', '').strip()

    # --- Validate required fields ---
    if not semester:
        return jsonify({"success": False, "error": "Semester (Program Name) is required"}), 400
    if not course_code:
        return jsonify({"success": False, "error": "Course Code is required"}), 400
    if not year:
        return jsonify({"success": False, "error": "Year is required"}), 400

    # --- 1. Resolve / auto-create Program ---
    prog = Program.query.filter_by(program_code=program_code).first() if program_code else None
    if not prog:
        # Create a minimal program entry so FK is satisfied
        prog = Program(
            program_code = program_code or semester,
            program_name = semester,
            department   = department
        )
        db.session.add(prog)
        db.session.flush()   # get prog.id without committing

    # --- 2. Resolve / auto-create Course (in the Course dim table) ---
    course = Course.query.filter_by(course_code=course_code).first()
    if not course:
        # Pull course name from the lookup table
        scl = SemesterCourseLookup.query.filter_by(
            semester=semester, course_code=course_code
        ).first()
        course_name = scl.course_name if scl else data.get('courseName', course_code)
        course = Course(course_code=course_code, course_name=course_name)
        db.session.add(course)
        db.session.flush()

    # --- 3. Duplicate check: same year + program + course ---
    existing = C11Courses.query.filter_by(
        academic_year=year,
        program_id=prog.id,
        course_id=course.id
    ).first()
    if existing:
        return jsonify({"success": False, "error": f"Course {course_code} for {semester} in {year} already exists"}), 400

    # --- 4. Insert record ---
    try:
        new_rec = C11Courses(
            academic_year = year,
            program_id    = prog.id,
            course_id     = course.id,
            proof_links   = data.get('proofLink'),
            created_by_id = session.get('user_id'),
            updated_by_id = session.get('user_id')
        )
        db.session.add(new_rec)
        db.session.commit()
        return jsonify({"success": True, "data": to_dict(new_rec)})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

@api_bp.route('/records/<criterion>', methods=['POST'])
def add_record(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify({"success": False, "error": "Criterion not found"}), 404
    
    data = request.json
    db_kwargs = {}
    
    # Audit tracking
    user_id = session.get('user_id')
    db_kwargs['created_by_id'] = user_id
    db_kwargs['updated_by_id'] = user_id
    
    # Select All Optimization
    if data.get('select_all') is True:
        if 'programCode' in data:
            prog = Program.query.filter_by(program_code=data['programCode']).first()
            all_studs = Student.query.filter_by(program_id=prog.id).all() if prog else Student.query.all()
        else:
            all_studs = Student.query.all()
        data['student_list'] = ", ".join([s.name for s in all_studs])

    # 1. Map common foreign keys dynamically based on UI selections
    if 'courseCode' in data:
        c = Course.query.filter_by(course_code=data['courseCode']).first()
        if c and hasattr(model, 'course_id'): db_kwargs['course_id'] = c.id
    if 'programCode' in data:
        p = Program.query.filter_by(program_code=data['programCode']).first()
        if p and hasattr(model, 'program_id'): db_kwargs['program_id'] = p.id
        
    # Resolve teacher_id or pi_id if model expects it
    if hasattr(model, 'teacher_id') or hasattr(model, 'pi_id'):
        t_id = resolve_teacher_id(model, data)
        if t_id:
            if hasattr(model, 'teacher_id'): db_kwargs['teacher_id'] = t_id
            if hasattr(model, 'pi_id'): db_kwargs['pi_id'] = t_id
        
    # Resolve student_id if model expects it
    if hasattr(model, 'student_id'):
        stud_id = resolve_student_id(model, data)
        if stud_id: db_kwargs['student_id'] = stud_id

    # 2. Map payload keys to model columns
    for col in model.__table__.columns:
        col_name = col.name
        if col_name == 'id': continue
        
        # Direct match or alias resolution
        val = None
        has_val = False
        if col_name in data:
            val = data[col_name]
            has_val = True
        else:
            # Common UI -> DB Alias mapping
            if col_name == 'academic_year' and 'year' in data:
                val = data['year']
                has_val = True
            elif col_name == 'year_of_offering' and 'year' in data:
                val = data['year']
                has_val = True
            elif col_name == 'year_of_passing' and 'year' in data:
                val = data['year']
                has_val = True
            elif col_name == 'year_of_implementation' and 'cbcsYear' in data:
                val = data['cbcsYear']
                has_val = True
            elif col_name == 'status_of_implementation' and 'cbcsStatus' in data:
                val = data['cbcsStatus']
                has_val = True
            elif col_name == 'institution_joined' and 'inst_joined' in data:
                val = data['inst_joined']
                has_val = True
            elif col_name == 'program_admitted' and 'prog_joined' in data:
                val = data['prog_joined']
                has_val = True
            elif col_name == 'program_graduated' and 'program' in data:
                val = data['program']
                has_val = True
            elif col_name == 'registration_number':
                val = data.get('registration_no') or data.get('reg_no') or data.get('enrollmentNumber') or data.get('registration_number')
                has_val = val is not None
            elif col_name == 'exam_type' and 'exam_qualified' in data:
                val = data['exam_qualified']
                has_val = True
            elif col_name == 'supporting_document' and 'upload_supporting_document' in data:
                val = data['upload_supporting_document']
                has_val = True
            elif col_name.startswith('earmarked_') and col_name.replace('earmarked_', 'ear_') in data:
                val = data[col_name.replace('earmarked_', 'ear_')]
                has_val = True
            elif col_name.startswith('admitted_') and col_name.replace('admitted_', 'adm_') in data:
                val = data[col_name.replace('admitted_', 'adm_')]
                has_val = True
            elif col_name == 'team_or_individual' and 'team_individual' in data:
                val = data['team_individual']
                has_val = True
            elif col_name == 'level' and 'event_level' in data:
                val = data['event_level']
                has_val = True
            elif col_name == 'activity_type' and 'event_name' in data:
                val = data['event_name']
                has_val = True
            # ---- Criteria 6 aliases ----
            elif col_name == 'area' and 'areas_of_e_governance' in data:
                val = data['areas_of_e_governance']
                has_val = True
            elif col_name == 'vendor_name' and 'name_of_vendor_with_contact_details' in data:
                val = data['name_of_vendor_with_contact_details']
                has_val = True
            elif col_name == 'conference_name' and 'name_of_conference_workshop' in data:
                val = data['name_of_conference_workshop']
                has_val = True
            elif col_name == 'professional_body' and 'name_of_professional_body' in data:
                val = data['name_of_professional_body']
                has_val = True
            elif col_name == 'amount' and 'amount_of_support' in data:
                val = data['amount_of_support']
                has_val = True
            elif col_name == 'teaching_program_title' and 'title_of_professional_development_program' in data:
                val = data['title_of_professional_development_program']
                has_val = True
            elif col_name == 'non_teaching_program_title' and 'title_of_administrative_training_program' in data:
                val = data['title_of_administrative_training_program']
                has_val = True
            elif col_name == 'participant_count' and 'no_of_participants' in data:
                val = data['no_of_participants']
                has_val = True
            elif col_name == 'program_title' and 'title_of_the_program' in data:
                val = data['title_of_the_program']
                has_val = True
            elif col_name == 'agency_name' and 'name_of_non_government_funding_agencies_individuals' in data:
                val = data['name_of_non_government_funding_agencies_individuals']
                has_val = True
            elif col_name == 'purpose' and 'purpose_of_the_grant' in data:
                val = data['purpose_of_the_grant']
                has_val = True
            elif col_name == 'amount_received' and 'funds_grants_received_inr_in_lakhs' in data:
                val = data['funds_grants_received_inr_in_lakhs']
                has_val = True
            elif col_name == 'audited_statement_link' and 'link_to_audited_statement' in data:
                val = data['link_to_audited_statement']
                has_val = True
            elif col_name == 'conferences_conducted' and 'conferences_seminars_workshops_on_quality' in data:
                val = data['conferences_seminars_workshops_on_quality']
                has_val = True
            elif col_name == 'aaa_status' and 'academic_administrative_audit_aaa' in data:
                val = data['academic_administrative_audit_aaa']
                has_val = True
            elif col_name == 'nirf_status' and 'participation_in_nirf' in data:
                val = data['participation_in_nirf']
                has_val = True
            elif col_name == 'nba_certification' and 'nba_or_other_certification' in data:
                val = data['nba_or_other_certification']
                has_val = True
            elif col_name == 'collaborative_initiatives' and 'collaborative_quality_initiatives' in data:
                val = data['collaborative_quality_initiatives']
                has_val = True
            elif col_name == 'orientation_program' and 'orientation_programme_on_quality_issues' in data:
                val = data['orientation_programme_on_quality_issues']
                has_val = True

        if has_val:
            if val is None or val == "":
                db_kwargs[col_name] = None
            elif col_name == 'status_of_implementation':
                db_kwargs[col_name] = str(val).lower() == "yes"
            elif isinstance(col.type, db.Date):
                if val:
                    try:
                        from datetime import datetime
                        db_kwargs[col_name] = datetime.strptime(str(val).split('T')[0], '%Y-%m-%d').date()
                    except Exception:
                        try:
                            db_kwargs[col_name] = datetime.strptime(str(val), '%d-%m-%Y').date()
                        except Exception:
                            db_kwargs[col_name] = None
                else:
                    db_kwargs[col_name] = None
            elif isinstance(col.type, (db.Integer, db.Numeric)):
                try:
                    db_kwargs[col_name] = int(val) if isinstance(col.type, db.Integer) else float(val)
                except ValueError:
                    import re
                    match = re.search(r'\d+(?:\.\d+)?', str(val))
                    if match:
                        try:
                            db_kwargs[col_name] = int(match.group()) if isinstance(col.type, db.Integer) else float(match.group())
                        except ValueError:
                            db_kwargs[col_name] = None
                    else:
                        db_kwargs[col_name] = None
            else:
                db_kwargs[col_name] = val

    # Default Academic Year from session if missing in payload
    if 'academic_year' not in db_kwargs and hasattr(model, 'academic_year'):
        db_kwargs['academic_year'] = session.get('academic_year')

    try:
        new_rec = model(**db_kwargs)
        db.session.add(new_rec)
        db.session.commit()
        return jsonify({"success": True, "data": to_dict(new_rec)})
    except Exception as e:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 400

@api_bp.route('/records/<criterion>/<int:id>', methods=['PUT'])
def update_record(criterion, id):
    model = CRITERIA_MODELS.get(criterion)
    if not model: return jsonify({"success": False}), 404
    rec = model.query.get(id)
    if not rec: return jsonify({"success": False}), 404
    
    data = request.json
    
    # Audit tracking
    rec.updated_by_id = session.get('user_id')
    
    # Select All Optimization
    if data.get('select_all') is True:
        if 'programCode' in data:
            prog = Program.query.filter_by(program_code=data['programCode']).first()
            all_studs = Student.query.filter_by(program_id=prog.id).all() if prog else Student.query.all()
        else:
            all_studs = Student.query.all()
        data['student_list'] = ", ".join([s.name for s in all_studs])

    if 'courseCode' in data:
        c = Course.query.filter_by(course_code=data['courseCode']).first()
        if c and hasattr(model, 'course_id'): rec.course_id = c.id
    if 'programCode' in data:
        p = Program.query.filter_by(program_code=data['programCode']).first()
        if p and hasattr(model, 'program_id'): rec.program_id = p.id
        
    # Resolve teacher_id or pi_id if model expects it
    if hasattr(model, 'teacher_id') or hasattr(model, 'pi_id'):
        t_id = resolve_teacher_id(model, data)
        if t_id:
            if hasattr(model, 'teacher_id'): rec.teacher_id = t_id
            if hasattr(model, 'pi_id'): rec.pi_id = t_id
        
    # Resolve student_id if model expects it
    if hasattr(model, 'student_id'):
        stud_id = resolve_student_id(model, data)
        if stud_id: rec.student_id = stud_id
        
    for col in model.__table__.columns:
        col_name = col.name
        if col_name == 'id': continue
        
        # Direct match or alias resolution
        val = None
        has_val = False
        if col_name in data:
            val = data[col_name]
            has_val = True
        else:
            # Common UI -> DB Alias mapping
            if col_name == 'academic_year' and 'year' in data:
                val = data['year']
                has_val = True
            elif col_name == 'year_of_offering' and 'year' in data:
                val = data['year']
                has_val = True
            elif col_name == 'year_of_passing' and 'year' in data:
                val = data['year']
                has_val = True
            elif col_name == 'year_of_implementation' and 'cbcsYear' in data:
                val = data['cbcsYear']
                has_val = True
            elif col_name == 'status_of_implementation' and 'cbcsStatus' in data:
                val = data['cbcsStatus']
                has_val = True
            elif col_name == 'institution_joined' and 'inst_joined' in data:
                val = data['inst_joined']
                has_val = True
            elif col_name == 'program_admitted' and 'prog_joined' in data:
                val = data['prog_joined']
                has_val = True
            elif col_name == 'program_graduated' and 'program' in data:
                val = data['program']
                has_val = True
            elif col_name == 'registration_number':
                val = data.get('registration_no') or data.get('reg_no') or data.get('enrollmentNumber') or data.get('registration_number')
                has_val = val is not None
            elif col_name == 'exam_type' and 'exam_qualified' in data:
                val = data['exam_qualified']
                has_val = True
            elif col_name == 'supporting_document' and 'upload_supporting_document' in data:
                val = data['upload_supporting_document']
                has_val = True
            elif col_name.startswith('earmarked_') and col_name.replace('earmarked_', 'ear_') in data:
                val = data[col_name.replace('earmarked_', 'ear_')]
                has_val = True
            elif col_name.startswith('admitted_') and col_name.replace('admitted_', 'adm_') in data:
                val = data[col_name.replace('admitted_', 'adm_')]
                has_val = True
            elif col_name == 'team_or_individual' and 'team_individual' in data:
                val = data['team_individual']
                has_val = True
            elif col_name == 'level' and 'event_level' in data:
                val = data['event_level']
                has_val = True
            elif col_name == 'activity_type' and 'event_name' in data:
                val = data['event_name']
                has_val = True
            # ---- Criteria 6 aliases ----
            elif col_name == 'area' and 'areas_of_e_governance' in data:
                val = data['areas_of_e_governance']
                has_val = True
            elif col_name == 'vendor_name' and 'name_of_vendor_with_contact_details' in data:
                val = data['name_of_vendor_with_contact_details']
                has_val = True
            elif col_name == 'conference_name' and 'name_of_conference_workshop' in data:
                val = data['name_of_conference_workshop']
                has_val = True
            elif col_name == 'professional_body' and 'name_of_professional_body' in data:
                val = data['name_of_professional_body']
                has_val = True
            elif col_name == 'amount' and 'amount_of_support' in data:
                val = data['amount_of_support']
                has_val = True
            elif col_name == 'teaching_program_title' and 'title_of_professional_development_program' in data:
                val = data['title_of_professional_development_program']
                has_val = True
            elif col_name == 'non_teaching_program_title' and 'title_of_administrative_training_program' in data:
                val = data['title_of_administrative_training_program']
                has_val = True
            elif col_name == 'participant_count' and 'no_of_participants' in data:
                val = data['no_of_participants']
                has_val = True
            elif col_name == 'program_title' and 'title_of_the_program' in data:
                val = data['title_of_the_program']
                has_val = True
            elif col_name == 'agency_name' and 'name_of_non_government_funding_agencies_individuals' in data:
                val = data['name_of_non_government_funding_agencies_individuals']
                has_val = True
            elif col_name == 'purpose' and 'purpose_of_the_grant' in data:
                val = data['purpose_of_the_grant']
                has_val = True
            elif col_name == 'amount_received' and 'funds_grants_received_inr_in_lakhs' in data:
                val = data['funds_grants_received_inr_in_lakhs']
                has_val = True
            elif col_name == 'audited_statement_link' and 'link_to_audited_statement' in data:
                val = data['link_to_audited_statement']
                has_val = True
            elif col_name == 'conferences_conducted' and 'conferences_seminars_workshops_on_quality' in data:
                val = data['conferences_seminars_workshops_on_quality']
                has_val = True
            elif col_name == 'aaa_status' and 'academic_administrative_audit_aaa' in data:
                val = data['academic_administrative_audit_aaa']
                has_val = True
            elif col_name == 'nirf_status' and 'participation_in_nirf' in data:
                val = data['participation_in_nirf']
                has_val = True
            elif col_name == 'nba_certification' and 'nba_or_other_certification' in data:
                val = data['nba_or_other_certification']
                has_val = True
            elif col_name == 'collaborative_initiatives' and 'collaborative_quality_initiatives' in data:
                val = data['collaborative_quality_initiatives']
                has_val = True
            elif col_name == 'orientation_program' and 'orientation_programme_on_quality_issues' in data:
                val = data['orientation_programme_on_quality_issues']
                has_val = True

        if has_val:
            if val is None or val == "":
                setattr(rec, col_name, None)
            elif col_name == 'status_of_implementation':
                setattr(rec, col_name, str(val).lower() == "yes")
            elif isinstance(col.type, db.Date):
                if val:
                    try:
                        from datetime import datetime
                        setattr(rec, col_name, datetime.strptime(str(val).split('T')[0], '%Y-%m-%d').date())
                    except Exception:
                        try:
                            setattr(rec, col_name, datetime.strptime(str(val), '%d-%m-%Y').date())
                        except Exception:
                            setattr(rec, col_name, None)
                else:
                    setattr(rec, col_name, None)
            elif isinstance(col.type, (db.Integer, db.Numeric)):
                try:
                    setattr(rec, col_name, int(val) if isinstance(col.type, db.Integer) else float(val))
                except ValueError:
                    import re
                    match = re.search(r'\d+(?:\.\d+)?', str(val))
                    if match:
                        try:
                            setattr(rec, col_name, int(match.group()) if isinstance(col.type, db.Integer) else float(match.group()))
                        except ValueError:
                            setattr(rec, col_name, None)
                    else:
                        setattr(rec, col_name, None)
            else:
                setattr(rec, col_name, val)
    db.session.commit()
    return jsonify({"success": True, "data": to_dict(rec)})

@api_bp.route('/records/<criterion>/<int:id>', methods=['DELETE'])
def delete_record(criterion, id):
    model = CRITERIA_MODELS.get(criterion)
    if model:
        rec = model.query.get(id)
        if rec:
            db.session.delete(rec)
            db.session.commit()
    return jsonify({"success": True})

@api_bp.route('/records/<criterion>/bulk-delete', methods=['POST'])
def bulk_delete(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if model:
        ids = request.json.get('ids', [])
        model.query.filter(model.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
    return jsonify({"success": True})

@api_bp.route('/upload-evidence', methods=['POST'])
def upload_evidence():
    criterion = request.form.get('criterion')
    record_id = request.form.get('record_id')
    files = request.files.getlist('files')
    
    if not criterion or not record_id or not files:
        return jsonify({"success": False, "error": "Missing criterion, record_id or files"}), 400
        
    model = CRITERIA_MODELS.get(criterion)
    if not model:
        return jsonify({"success": False, "error": "Invalid criterion"}), 400
        
    rec = model.query.get(record_id)
    if not rec:
        return jsonify({"success": False, "error": "Record not found"}), 404
        
    upload_folder = current_app.config.get('UPLOAD_FOLDER', 'static/uploads')
    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)
        
    file_urls = []
    for f in files:
        if f.filename:
            filename = secure_filename(f.filename)
            unique_name = f"{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(upload_folder, unique_name)
            f.save(file_path)
            
            ev = Evidence(
                filename=filename,
                file_path=f"/{upload_folder}/{unique_name}",
                criterion_type=criterion,
                record_id=record_id,
                created_by_id=session.get('user_id')
            )
            db.session.add(ev)
            file_urls.append(f"/{upload_folder}/{unique_name}")
            
    if file_urls:
        existing_links = rec.proof_links.split(',') if getattr(rec, 'proof_links', None) else []
        existing_links.extend(file_urls)
        existing_links = [l.strip() for l in existing_links if l.strip()]
        rec.proof_links = ','.join(existing_links)
        
    db.session.commit()
    return jsonify({"success": True, "file_urls": file_urls})

# --- Excel Export Endpoint ---

COLUMN_HEADER_MAP = {
    # General Audit
    "id": "ID",
    "year": "Year",
    "academicYear": "Academic Year",
    "createdBy": "Created By",
    "updatedBy": "Updated By",
    "created_at": "Created At",
    "updated_at": "Updated At",
    
    # 1.1
    "programName": "Program Name",
    "courseCode": "Course Code",
    "courseName": "Course Name",
    
    # 1.1.3 / 1.3.2 / 1.3.3 / etc.
    "teacherName": "Teacher Name",
    "department": "Department",
    "program": "Program Name",
    "program_graduated": "Program Graduated",
    "employer_name": "Employer Name",
    "employer_contact": "Employer Contact",
    "pay_package": "Pay Package (LPA)",
    "inst_joined": "Institution Joined",
    "prog_joined": "Program Admitted To",
    "registration_no": "Registration No / Roll No",
    "exam_qualified": "Exam Qualified",
    "studentName": "Student Name",
    "student_name": "Student Name",
    "enrollmentNumber": "Enrollment Number",
    "category": "Category",
    "exam_activity_name": "Exam Activity Name",
    "exam_students_count": "Number of Students (Exams)",
    "counseling_activity_name": "Counseling Activity Name",
    "counseling_students_count": "Number of Students (Counseling)",
    "students_placed": "Number of Students Placed",
    
    # Other common fields
    "achievement": "Achievement",
    "award_name": "Award Name",
    "award_category": "Award Category",
    "level": "Level",
    "event_name": "Event Name",
    "team_individual": "Team/Individual",
    "organization": "Organization",
    "position": "Position",
    "nature_of_appointment": "Nature of Appointment",
    "pan": "PAN",
    "designation": "Designation",
    "joining_date": "Joining Date",
    "leaving_date": "Leaving Date",
    "highest_degree": "Highest Degree",
    "degree_year": "Degree Year",
    "proof_links": "Proof Link(s)",
    "pdf_path": "PDF Path",
}

def make_header_nice(key):
    if key in COLUMN_HEADER_MAP:
        return COLUMN_HEADER_MAP[key]
    import re
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1 \2', key)
    s2 = re.sub('([a-z0-9])([A-Z])', r'\1 \2', s1)
    return s2.replace('_', ' ').title()

@api_bp.route('/export-excel/<criterion>', methods=['GET'])
def export_excel(criterion):
    model = CRITERIA_MODELS.get(criterion)
    if not model:
        return jsonify({"success": False, "error": "Invalid criterion"}), 400

    # Retrieve all records
    if criterion == '1_1':
        records = C11Courses.query.all()
        result = []
        for r in records:
            d = to_dict(r)
            if r.course_id:
                c = Course.query.get(r.course_id)
                if c:
                    scl = SemesterCourseLookup.query.filter_by(course_code=c.course_code).first()
                    if scl: 
                        d['programName'] = scl.semester
            result.append(d)
    else:
        result = [to_dict(r) for r in model.query.all()]

    EXPORT_HEADERS = {
        "5_3_1": {
            "year": "Year",
            "award_name": "Name of the award/medal",
            "team_or_individual": "Team / Individual",
            "level": "University/State/National/International",
            "activity_type": "Sports/Cultural",
            "student_name": "Name of the student"
        },
        "5_3_3": {
            "event_date": "Date of event/activity (DD-MM-YYYY)",
            "event_name": "Name of the event/activity",
            "student_name": "Name of the student participated"
        }
    }

    if not result:
        df = pd.DataFrame()
    elif criterion in EXPORT_HEADERS:
        headers_map = EXPORT_HEADERS[criterion]
        filtered_result = []
        for r in result:
            new_row = {}
            for col_key, header_name in headers_map.items():
                val = r.get(col_key)
                if col_key == "event_date" and val:
                    try:
                        parsed_date = datetime.strptime(str(val).split('T')[0], '%Y-%m-%d')
                        new_row[header_name] = parsed_date.strftime('%d-%m-%Y')
                    except Exception:
                        try:
                            parsed_date = datetime.strptime(str(val), '%d-%m-%Y')
                            new_row[header_name] = parsed_date.strftime('%d-%m-%Y')
                        except Exception:
                            new_row[header_name] = val
                else:
                    new_row[header_name] = val if val is not None else ""
            filtered_result.append(new_row)
        df = pd.DataFrame(filtered_result)
    else:
        df = pd.DataFrame(result)
        cols = list(df.columns)
        first_cols = [c for c in ['id', 'year', 'academicYear', 'studentName', 'student_name', 'enrollmentNumber'] if c in cols]
        other_cols = [c for c in cols if c not in first_cols]
        new_order = first_cols + other_cols
        df = df[new_order]
        df.rename(columns=lambda x: make_header_nice(x), inplace=True)

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Criterion_{criterion}"

    title_text = f"Criteria {criterion.replace('_', '.')}"
    
    # We style up to max 15 columns for visual cleanliness
    max_cols = max(len(df.columns), 8)
    max_col_letter = get_column_letter(max_cols)

    ws.merge_cells(f'A1:{max_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.font = Font(bold=True, size=14, color='FFFFFF', name='Arial')
    title_cell.fill = PatternFill('solid', start_color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells(f'A2:{max_col_letter}2')
    inst_cell = ws['A2']
    inst_cell.value = 'MET Bhujbal Knowledge City – Institute of Engineering, Nashik'
    inst_cell.font = Font(bold=True, size=10, name='Arial', italic=True)
    inst_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    header_fill = PatternFill('solid', start_color='2C3E50')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    data_font = Font(name='Arial', size=10)
    for row_idx, row_data in enumerate(df.values, start=4):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is None:
                cell.value = "—"
            elif isinstance(val, (datetime, date)):
                cell.value = val.strftime('%Y-%m-%d')
            else:
                cell.value = val

            cell.font = data_font
            cell.border = thin_border
            
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row_idx].height = 20

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    if len(ws.views.sheetView) > 0:
        ws.views.sheetView[0].showGridLines = True
    else:
        # If views are empty, create one
        ws.views.sheetView.append(openpyxl.worksheet.views.SheetView())
        ws.views.sheetView[0].showGridLines = True

    wb.save(output)
    output.seek(0)

    filename = f"Criterion_{criterion}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

